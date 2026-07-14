import json
import re
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


ROOT = Path(__file__).resolve().parent
PROCESSING_DIR = ROOT / "processing"
OUTPUT_DIR = ROOT / "output"

TITLE_ROW = 1
SUBTITLE_ROW = 2
COST_NAME_ROW = 3
APPLY_IF_ROW = 4
VALIDITY_ROW = 5
PROLONG_ROW = 6
RATE_BY_ROW = 7
RULE_ROW = 8
HEADER_ROW = 10
DATA_START_ROW = 11

GREEN_FILL = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")
HEADER_FILL = PatternFill(fill_type="solid", start_color="D9E1F2", end_color="D9E1F2")
META_FILL = PatternFill(fill_type="solid", start_color="F2F2F2", end_color="F2F2F2")


def choose_updated_json():
    files = sorted(PROCESSING_DIR.glob("*updated*.json"))
    if not files:
        raise FileNotFoundError(f"No *_updated.json files found in {PROCESSING_DIR}")

    if len(sys.argv) > 1:
        arg = sys.argv[1].strip()
        if arg.isdigit():
            idx = int(arg) - 1
            if 0 <= idx < len(files):
                return files[idx]
            raise ValueError(f"Index out of range: {arg}")
        candidate = PROCESSING_DIR / arg
        if candidate.exists():
            return candidate
        raise FileNotFoundError(f"Could not find file in processing: {arg}")

    print("Choose updated JSON file:")
    for i, path in enumerate(files, start=1):
        print(f"{i}. {path.name}")
    selected = input("Enter file number: ").strip()
    if not selected.isdigit():
        raise ValueError("Please enter a valid number.")
    idx = int(selected) - 1
    if not (0 <= idx < len(files)):
        raise ValueError("Selected number is out of range.")
    return files[idx]


def collect_route_keys(records):
    return list(records[0].get("route", {}).keys()) if records else []


def parse_ddmmyyyy(value):
    if not value:
        return None
    try:
        return datetime.strptime(str(value), "%d.%m.%Y").date()
    except ValueError:
        return None


def parse_period(value):
    if not value:
        return (None, None)
    raw = str(value).strip()
    if "-" not in raw:
        return (None, None)
    a, b = raw.split("-", 1)
    start = parse_ddmmyyyy(a.strip())
    end = parse_ddmmyyyy(b.strip()) if b.strip().upper() != "N/A" else None
    return (start, end)


def is_cost_within_rate_card(cost_validity, card_validity):
    if not card_validity:
        return True
    cost_start, cost_end = parse_period(cost_validity)
    card_start, card_end = parse_period(card_validity)
    if not (card_start and card_end):
        return True
    if not (cost_start and cost_end):
        # Costs without explicit validity remain visible.
        return True
    return not (cost_end < card_start or cost_start > card_end)


def collect_cost_order(records, payload):
    card_validity = payload.get("rate_card_validity")
    seen = set()
    cost_meta_for_sort = {}
    first_seen = []
    for rec in records:
        for rate in rec.get("rates", []):
            name = str(rate.get("cost_name", ""))
            if not name:
                continue
            if not is_cost_within_rate_card(rate.get("validity_period"), card_validity):
                continue
            if name not in seen:
                seen.add(name)
                first_seen.append(name)
            if name not in cost_meta_for_sort:
                cost_meta_for_sort[name] = {
                    "container": str(rate.get("container_type") or ""),
                    "validity": str(rate.get("validity_period") or ""),
                }

    def container_key(container):
        order = {"22G0": 22, "25G0": 25, "42G0": 42, "45G0": 45, "52G0": 52}
        return order.get(container, 999)

    def group_key(cost_name):
        if cost_name.startswith("BAF Fee"):
            return 0
        if cost_name.startswith("EU ETS Fee"):
            return 1
        return 2

    def validity_key(validity):
        start, _ = parse_period(validity)
        if start is None:
            return 99999999
        return int(start.strftime("%Y%m%d"))

    def is_plain_baf_ets(cost_name):
        name = str(cost_name)
        if name in ("BAF Fee", "EU ETS Fee"):
            return True
        # Plain form has only container in parentheses, e.g. "BAF Fee (22G0)".
        return bool(re.match(r"^(BAF Fee|EU ETS Fee) \([A-Z0-9]+\)$", name))

    def container_key_for_sort(cost_name, container):
        name = str(cost_name)
        if name in ("BAF Fee", "EU ETS Fee"):
            return -1
        order = {"22G0": 22, "25G0": 25, "42G0": 42, "45G0": 45, "52G0": 52}
        return order.get(str(container), 999)

    # Apply container/group sorting ONLY to BAF/EU ETS, keep others as-is.
    special = [n for n in first_seen if group_key(n) in (0, 1)]
    special_sorted = sorted(
        special,
        key=lambda n: (
            container_key_for_sort(n, cost_meta_for_sort[n]["container"]),
            group_key(n),
            0 if is_plain_baf_ets(n) else 1,
            validity_key(cost_meta_for_sort[n]["validity"]),
            n,
        ),
    )

    # Final order:
    # 1) all non-BAF/EU ETS in original order
    # 2) BAF/EU ETS in sorted order
    # Non-special columns should follow the previous rate-card order.
    # The first lane preserves that order best, so use it as the baseline.
    baseline = []
    if records:
        for rate in records[0].get("rates", []):
            name = str(rate.get("cost_name", ""))
            if name and name in seen and group_key(name) not in (0, 1) and name not in baseline:
                baseline.append(name)
    # Add any remaining non-special costs that did not appear in first lane.
    non_special = baseline + [n for n in first_seen if group_key(n) not in (0, 1) and n not in baseline]
    ordered = non_special + special_sorted
    return ordered


def collect_cost_meta(records, payload):
    forced_new_costs = set(payload.get("update_context", {}).get("new_cost_names", []))
    meta = {}
    for rec in records:
        for rate in rec.get("rates", []):
            name = str(rate.get("cost_name", ""))
            if not name:
                continue
            if name not in meta:
                meta[name] = {
                    "apply_if": rate.get("apply_if"),
                    "validity_period": rate.get("validity_period"),
                    "cost_to_prolong": rate.get("cost_to_prolong"),
                    "rate_by": rate.get("rate_by"),
                    "rule": rate.get("rule"),
                    "is_new": (rate.get("update_note") == "(new)") or (name in forced_new_costs),
                    "is_changed": rate.get("update_note") in {"(new)", "(updated)"},
                }
            else:
                for k in ("apply_if", "validity_period", "cost_to_prolong", "rate_by", "rule"):
                    if not meta[name].get(k) and rate.get(k):
                        meta[name][k] = rate.get(k)
                if rate.get("update_note") == "(new)" or name in forced_new_costs:
                    meta[name]["is_new"] = True
                if rate.get("update_note") in {"(new)", "(updated)"}:
                    meta[name]["is_changed"] = True
    return meta


def should_show_flat_column(cost_name):
    name = str(cost_name)
    return name == "Base Rate (BASE_STAT_FRK)" or name == "RoRo Fee"


def collect_cost_metrics(records, cost_names):
    metrics = {}
    for cost_name in cost_names:
        if should_show_flat_column(cost_name):
            metrics[cost_name] = ["currency", "flat_min", "p_unit"]
        else:
            metrics[cost_name] = ["currency", "p_unit"]
    return metrics


def metric_label(metric_key):
    return "Currency" if metric_key == "currency" else ("Flat" if metric_key == "flat_min" else "p/unit")


def write_layout(ws, payload, route_keys, cost_names, cost_meta, cost_metrics):
    ws.title = "Rate card"
    ws.cell(TITLE_ROW, 1, "Rate Card (Generated from JSON)")
    ws.cell(TITLE_ROW, 1).font = Font(bold=True, size=12)
    ws.cell(SUBTITLE_ROW, 1, f"Source JSON: {payload.get('source_file', '')}")
    ws.cell(SUBTITLE_ROW, 2, f"Record count: {payload.get('record_count', 0)}")

    col = 1
    for route_name in route_keys:
        ws.cell(HEADER_ROW, col, route_name)
        ws.cell(HEADER_ROW, col).font = Font(bold=True)
        ws.cell(HEADER_ROW, col).fill = HEADER_FILL
        ws.column_dimensions[ws.cell(HEADER_ROW, col).column_letter].width = 18
        col += 1

    for cost_name in cost_names:
        metric_keys = cost_metrics[cost_name]
        span = len(metric_keys)
        ws.merge_cells(start_row=COST_NAME_ROW, start_column=col, end_row=COST_NAME_ROW, end_column=col + span - 1)
        ws.cell(COST_NAME_ROW, col, cost_name)
        ws.cell(COST_NAME_ROW, col).font = Font(bold=True)

        meta = cost_meta.get(cost_name, {})
        apply_if = meta.get("apply_if")
        validity = meta.get("validity_period")
        prolong = meta.get("cost_to_prolong")
        rate_by = meta.get("rate_by")
        rule = meta.get("rule")
        is_changed_cost = bool(meta.get("is_changed"))
        meta_fill = GREEN_FILL if is_changed_cost else META_FILL
        ws.cell(COST_NAME_ROW, col).fill = meta_fill

        ws.merge_cells(start_row=APPLY_IF_ROW, start_column=col, end_row=APPLY_IF_ROW, end_column=col + span - 1)
        ws.cell(APPLY_IF_ROW, col, f"Apply if: {apply_if}" if apply_if else None)
        ws.cell(APPLY_IF_ROW, col).fill = meta_fill

        ws.merge_cells(start_row=VALIDITY_ROW, start_column=col, end_row=VALIDITY_ROW, end_column=col + span - 1)
        ws.cell(VALIDITY_ROW, col, f"Validity: {validity}" if validity else None)
        ws.cell(VALIDITY_ROW, col).fill = meta_fill

        ws.merge_cells(start_row=PROLONG_ROW, start_column=col, end_row=PROLONG_ROW, end_column=col + span - 1)
        ws.cell(PROLONG_ROW, col, f"Prolong: {prolong}" if prolong else None)
        ws.cell(PROLONG_ROW, col).fill = meta_fill

        ws.merge_cells(start_row=RATE_BY_ROW, start_column=col, end_row=RATE_BY_ROW, end_column=col + span - 1)
        ws.cell(RATE_BY_ROW, col, f"Rate by: {rate_by}" if rate_by else None)
        ws.cell(RATE_BY_ROW, col).fill = meta_fill

        ws.merge_cells(start_row=RULE_ROW, start_column=col, end_row=RULE_ROW, end_column=col + span - 1)
        ws.cell(RULE_ROW, col, f"Rule: {rule}" if rule else None)
        ws.cell(RULE_ROW, col).fill = meta_fill

        for offset, metric_key in enumerate(metric_keys):
            c = col + offset
            ws.cell(HEADER_ROW, c, metric_label(metric_key))
            ws.cell(HEADER_ROW, c).font = Font(bold=True)
            ws.cell(HEADER_ROW, c).fill = HEADER_FILL
            ws.column_dimensions[ws.cell(HEADER_ROW, c).column_letter].width = 14
        col += span


def write_records(ws, records, route_keys, cost_names, cost_metrics, payload):
    route_count = len(route_keys)
    cost_start_col = {}
    running = route_count + 1
    for name in cost_names:
        cost_start_col[name] = running
        running += len(cost_metrics[name])

    sheet_used = str(payload.get("update_context", {}).get("sheet_used", "")).upper()
    highlight_route = sheet_used not in {"ETSBAF", "EU ETS", "EU_ETS"}

    for i, rec in enumerate(records):
        row = DATA_START_ROW + i
        route = rec.get("route", {})
        route_updated = route.get("update_note") in {"(new)", "(updated)"}
        changed_fields = set(route.get("update_changed_fields", []))
        route_source = str(route.get("update_source", "")).upper()

        for j, key in enumerate(route_keys):
            col = j + 1
            ws.cell(row, col, route.get(key))
            if not highlight_route and route_source != "BASE" and not changed_fields:
                continue
            if key in changed_fields:
                ws.cell(row, col).fill = GREEN_FILL
            elif route_updated and not changed_fields:
                ws.cell(row, col).fill = GREEN_FILL

        for rate in rec.get("rates", []):
            name = str(rate.get("cost_name", ""))
            if name not in cost_start_col:
                continue
            base_col = cost_start_col[name]
            metric_keys = cost_metrics[name]
            for offset, metric_key in enumerate(metric_keys):
                # Layout rule: keep Flat (MIN) column but never populate values.
                value = None if metric_key == "flat_min" else rate.get(metric_key)
                ws.cell(row, base_col + offset, value)
            if rate.get("update_note") in {"(new)", "(updated)"}:
                for offset in range(len(metric_keys)):
                    ws.cell(row, base_col + offset).fill = GREEN_FILL


def main():
    json_path = choose_updated_json()
    payload = json.loads(json_path.read_text(encoding="utf-8"))
    records = payload.get("records", [])
    if not records:
        raise ValueError("No records found in updated JSON.")

    route_keys = collect_route_keys(records)
    cost_names = collect_cost_order(records, payload)
    cost_meta = collect_cost_meta(records, payload)
    cost_metrics = collect_cost_metrics(records, cost_names)

    wb = Workbook()
    ws = wb.active
    write_layout(ws, payload, route_keys, cost_names, cost_meta, cost_metrics)
    write_records(ws, records, route_keys, cost_names, cost_metrics, payload)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_xlsx = OUTPUT_DIR / f"{json_path.stem}.xlsx"
    try:
        wb.save(output_xlsx)
    except PermissionError:
        fallback = OUTPUT_DIR / f"{json_path.stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(fallback)
        output_xlsx = fallback
    print(f"Created: {output_xlsx}")
    print("Template used: None (JSON-only layout)")


if __name__ == "__main__":
    main()
