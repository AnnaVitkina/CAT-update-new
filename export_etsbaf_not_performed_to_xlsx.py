import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from paths import OUTPUT_DIR, PROCESSING_DIR

HEADER_FILL = PatternFill(fill_type="solid", start_color="D9E1F2", end_color="D9E1F2")
HEADER_FONT = Font(bold=True)

COLUMNS = [
    ("KEY", lambda item: item.get("update_row", {}).get("route", {}).get("KEY")),
    ("Reason", lambda item: item.get("reason")),
    ("Transporeon ID", lambda item: item.get("transporeon_id")),
    ("Sheet", lambda item: item.get("update_row", {}).get("sheet_name")),
    ("Row #", lambda item: item.get("update_row", {}).get("row_number")),
    ("Carrier", lambda item: item.get("update_row", {}).get("route", {}).get("CARRIER")),
    ("Service", lambda item: item.get("update_row", {}).get("route", {}).get("SERVICE__C")),
    ("Valid from", lambda item: item.get("update_row", {}).get("route", {}).get("RATE_EFFECTIVE_DATE__C")),
    ("Valid to", lambda item: item.get("update_row", {}).get("route", {}).get("RATE_EXPIRATION_DATE__C")),
    ("Origin", lambda item: item.get("update_row", {}).get("route", {}).get("ORIGIN_LOCATION_NAME__C")),
    ("Destination", lambda item: item.get("update_row", {}).get("route", {}).get("DESTINATION_LOCATION_NAME__C")),
    ("Charge codes", lambda item: _charge_codes(item)),
]


def _charge_codes(item):
    rates = item.get("update_row", {}).get("rates", [])
    parts = []
    for rate in rates:
        charge = str(rate.get("charge_code") or "").upper()
        container = str(rate.get("container_type") or "")
        if charge and container:
            parts.append(f"{charge} ({container})")
        elif charge:
            parts.append(charge)
    return ", ".join(parts)


def choose_not_performed_json():
    files = sorted(PROCESSING_DIR.glob("*_etsbaf_not_performed.json"))
    if not files:
        raise FileNotFoundError(f"No *_etsbaf_not_performed.json files found in {PROCESSING_DIR}")

    if len(sys.argv) > 1:
        arg = sys.argv[1].strip()
        if arg.isdigit():
            idx = int(arg) - 1
            if 0 <= idx < len(files):
                return files[idx]
            raise ValueError(f"Index out of range: {arg}")
        candidate = PROCESSING_DIR / arg
        if candidate.exists():
            return candidate
        raise FileNotFoundError(f"Could not find not-performed JSON: {arg}")

    print("Choose ETSBAF not-performed JSON file:")
    for i, file_path in enumerate(files, start=1):
        print(f"{i}. {file_path.name}")
    selected = input("Enter file number: ").strip()
    if not selected.isdigit():
        raise ValueError("Please enter a valid number.")
    idx = int(selected) - 1
    if not (0 <= idx < len(files)):
        raise ValueError("Selected number is out of range.")
    return files[idx]


def output_path_for(json_path: Path) -> Path:
    return OUTPUT_DIR / f"{json_path.stem}.xlsx"


def autosize_columns(ws, num_cols, num_rows):
    for col_idx in range(1, num_cols + 1):
        letter = get_column_letter(col_idx)
        max_len = len(str(ws.cell(row=1, column=col_idx).value or ""))
        for row_idx in range(2, num_rows + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is not None:
                max_len = max(max_len, len(str(value)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 60)


def export_to_xlsx(json_path: Path, output_xlsx: Path) -> int:
    payload = json.loads(json_path.read_text(encoding="utf-8"))
    items = payload.get("items", [])

    wb = Workbook()
    ws = wb.active
    ws.title = "ETSBAF not performed"

    for col_idx, (header, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for row_idx, item in enumerate(items, start=2):
        for col_idx, (_, getter) in enumerate(COLUMNS, start=1):
            value = getter(item)
            ws.cell(row=row_idx, column=col_idx, value=value)

    autosize_columns(ws, len(COLUMNS), len(items) + 1)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)
    return len(items)


def main():
    json_path = choose_not_performed_json()
    output_xlsx = output_path_for(json_path)
    count = export_to_xlsx(json_path, output_xlsx)
    print(f"Wrote {count} rows to {output_xlsx}")
    print(f"Source JSON: {json_path}")


if __name__ == "__main__":
    main()
