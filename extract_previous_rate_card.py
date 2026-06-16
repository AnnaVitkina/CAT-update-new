import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
INPUT_DIR = ROOT / "input" / "previous rate card"
OUTPUT_DIR = ROOT / "processing"
TARGET_SHEET = "Rate card"


def to_primitive(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    return value


def safe_float(value):
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def clean_text(value):
    if value is None:
        return None
    text = str(value).replace("_x000D_", "\n")
    text = re.sub(r"\r\n?", "\n", text)
    text = re.sub(r"\n{2,}", "\n", text)
    text = text.strip()
    return text or None


def normalize_charge_code(title: str):
    title_no_paren = re.sub(r"\([^)]*\)", "", title).strip()
    code = re.sub(r"[^a-z0-9]+", "_", title_no_paren.lower()).strip("_")
    return code or "unknown_charge"


def extract_container_type(title: str):
    match = re.search(r"\(([^)]*)\)", title or "")
    if not match:
        return None
    inside = match.group(1).strip()
    container_match = re.search(r"\b([0-9]{2}G[0-9A-Z])\b", inside)
    if container_match:
        return container_match.group(1)
    return None


def normalize_field_name(name: str):
    normalized = re.sub(r"[^a-z0-9]+", "_", (name or "").strip().lower()).strip("_")
    return normalized or "unnamed"


def build_column_keys(data_header, charge_title_row):
    keys = []
    used = {}
    for col_idx in range(len(data_header)):
        metric_label = (
            str(data_header[col_idx]).strip() if data_header[col_idx] is not None else ""
        )
        charge_title = (
            str(charge_title_row[col_idx]).strip()
            if charge_title_row[col_idx] is not None
            else ""
        )

        if metric_label in {"Currency", "Flat", "p/unit"} and charge_title:
            metric_suffix = (
                "currency" if metric_label == "Currency" else ("flat" if metric_label == "Flat" else "p_unit")
            )
            base = f"{normalize_charge_code(charge_title)}_{metric_suffix}"
        else:
            base = normalize_field_name(metric_label)
            if base == "unnamed":
                base = f"column_{col_idx + 1}"

        count = used.get(base, 0) + 1
        used[base] = count
        keys.append(base if count == 1 else f"{base}_{count}")
    return keys


def extract_from_text(pattern, text):
    if not text:
        return None
    match = re.search(pattern, text, flags=re.IGNORECASE)
    if not match:
        return None
    if match.lastindex and match.lastindex >= 1:
        return match.group(1).strip()
    return match.group(0).strip()


def normalize_validity_period(value):
    if not value:
        return None
    full = value.strip()
    range_match = re.search(
        r"from\s+(\d{2}\.\d{2}\.\d{4})\s+to\s+(\d{2}\.\d{2}\.\d{4})",
        full,
        flags=re.IGNORECASE,
    )
    if range_match:
        return f"{range_match.group(1)}-{range_match.group(2)}"

    start_only = re.search(r"from\s+(\d{2}\.\d{2}\.\d{4})", full, flags=re.IGNORECASE)
    if start_only:
        return f"{start_only.group(1)}-N/A"
    return full


def extract_rate_card_validity(workbook):
    if "General info" not in workbook.sheetnames:
        return None
    ws = workbook["General info"]
    for row in ws.iter_rows(min_row=1, max_row=50, min_col=1, max_col=3, values_only=True):
        label = str(row[0]).strip() if row and row[0] is not None else ""
        if label.lower() == "validity period":
            raw = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            match = re.search(r"(\d{2}\.\d{2}\.\d{4})\s*-\s*(\d{2}\.\d{2}\.\d{4})", raw)
            if match:
                return f"{match.group(1)}-{match.group(2)}"
            return raw or None
    return None


def list_input_files():
    files = sorted(INPUT_DIR.glob("*.xlsx"))
    if not files:
        raise FileNotFoundError(f"No .xlsx files found in {INPUT_DIR}")
    return files


def choose_input_file(files):
    if len(sys.argv) > 1:
        arg = sys.argv[1].strip()
        if arg.isdigit():
            idx = int(arg) - 1
            if 0 <= idx < len(files):
                return files[idx]
            raise ValueError(f"File index out of range: {arg}")
        candidate = INPUT_DIR / arg
        if candidate.exists():
            return candidate
        raise FileNotFoundError(f"Could not find file in input folder: {arg}")

    print("Choose a previous rate card file to process:")
    for i, file_path in enumerate(files, start=1):
        print(f"{i}. {file_path.name}")

    selected = input("Enter file number: ").strip()
    if not selected.isdigit():
        raise ValueError("Please enter a valid number.")
    idx = int(selected) - 1
    if not (0 <= idx < len(files)):
        raise ValueError("Selected number is out of range.")
    return files[idx]


def find_header_row(rows):
    for idx, row in enumerate(rows):
        first_cell = str(row[0]).strip() if row and row[0] is not None else ""
        if first_cell == "Lane #":
            return idx
    raise ValueError("Could not find the 'Lane #' header row in 'Rate card' sheet.")


def parse_file(file_path: Path):
    workbook = load_workbook(file_path, data_only=True)
    if TARGET_SHEET not in workbook.sheetnames:
        raise ValueError(f"Sheet '{TARGET_SHEET}' not found in {file_path.name}")
    rate_card_validity = extract_rate_card_validity(workbook)

    ws = workbook[TARGET_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header_idx = find_header_row(rows)
    data_header = rows[header_idx]
    charge_title_row = rows[header_idx - 4] if header_idx >= 4 else tuple([None] * len(data_header))
    apply_if_row = rows[header_idx - 3] if header_idx >= 3 else tuple([None] * len(data_header))
    rate_by_row = rows[header_idx - 2] if header_idx >= 2 else tuple([None] * len(data_header))
    rule_row = rows[header_idx - 1] if header_idx >= 1 else tuple([None] * len(data_header))

    charges_by_column = {}
    ordered_cost_keys = []
    current_charge_title = None
    for col_idx in range(len(data_header)):
        metric_label = str(data_header[col_idx]).strip() if data_header[col_idx] is not None else ""
        charge_title = str(charge_title_row[col_idx]).strip() if charge_title_row[col_idx] is not None else ""
        if metric_label not in {"Currency", "Flat", "p/unit"}:
            continue
        if charge_title:
            current_charge_title = charge_title
        if not current_charge_title:
            continue

        metric = "currency" if metric_label == "Currency" else ("min" if metric_label == "Flat" else "rate")
        apply_if_text = clean_text(apply_if_row[col_idx]) if col_idx < len(apply_if_row) else None
        rate_by_text = clean_text(rate_by_row[col_idx]) if col_idx < len(rate_by_row) else None
        rule_text = clean_text(rule_row[col_idx]) if col_idx < len(rule_row) else None

        rate_by_value = extract_from_text(r"Rate by:\s*([^\n]+)", rate_by_text) or rate_by_text
        rule_value = (
            extract_from_text(r"rule:\s*([^\n]+)", rate_by_text)
            or extract_from_text(r"Regular rule|MIN|MAX", rate_by_text)
            or rule_text
        )
        charges_by_column[col_idx] = {
            "charge_code": normalize_charge_code(current_charge_title),
            "container_type": extract_container_type(current_charge_title),
            "cost_name": current_charge_title,
            "metric": metric,
            "apply_if": extract_from_text(r"(Applies if[^\n]+)", apply_if_text),
            "validity_period": normalize_validity_period(
                extract_from_text(r"Validity period:\s*([^\n]+)", apply_if_text)
            ),
            "cost_to_prolong": extract_from_text(r"Cost to prolong:\s*([^\n]+)", apply_if_text),
            "rate_by": rate_by_value,
            "rule": rule_value,
        }
        cost_key = (
            current_charge_title,
            extract_container_type(current_charge_title),
            normalize_charge_code(current_charge_title),
        )
        if cost_key not in ordered_cost_keys:
            ordered_cost_keys.append(cost_key)

    records = []
    for row_idx in range(header_idx + 1, len(rows)):
        row = rows[row_idx]
        if all(cell is None or cell == "" for cell in row):
            continue

        key = to_primitive(row[2]) if len(row) > 2 else None
        if key in (None, ""):
            continue

        record = {
            "row_number": row_idx + 1,
            "route": {
                "Lane #": to_primitive(row[0]) if len(row) > 0 else None,
                "Transporeon ID": to_primitive(row[1]) if len(row) > 1 else None,
                "KEY": key,
                "Carrier": to_primitive(row[3]) if len(row) > 3 else None,
                "SERVICE": to_primitive(row[4]) if len(row) > 4 else None,
                "SERVICE_C": to_primitive(row[5]) if len(row) > 5 else None,
                "Service": to_primitive(row[6]) if len(row) > 6 else None,
                "Valid from": to_primitive(row[7]) if len(row) > 7 else None,
                "Valid to": to_primitive(row[8]) if len(row) > 8 else None,
                "Origin Port": to_primitive(row[9]) if len(row) > 9 else None,
                "ORIGIN_COUNTRY__C": to_primitive(row[10]) if len(row) > 10 else None,
                "Destination Port": to_primitive(row[11]) if len(row) > 11 else None,
                "DESTINATION_COUNTRY__C": to_primitive(row[12]) if len(row) > 12 else None,
            },
            "rates": [],
        }

        # Initialize all known costs so each row always has a complete rate list.
        grouped = {}
        for group_key in ordered_cost_keys:
            matching_spec = next(
                (
                    spec
                    for spec in charges_by_column.values()
                    if (spec["cost_name"], spec["container_type"], spec["charge_code"]) == group_key
                ),
                None,
            )
            if matching_spec is None:
                continue
            if group_key not in grouped:
                grouped[group_key] = {
                    "cost_name": matching_spec["cost_name"],
                    "container_type": matching_spec["container_type"],
                    "apply_if": matching_spec["apply_if"],
                    "validity_period": matching_spec["validity_period"],
                    "cost_to_prolong": matching_spec["cost_to_prolong"],
                    "rate_by": matching_spec["rate_by"],
                    "rule": matching_spec["rule"],
                    "currency": None,
                    "flat_min": None,
                    "p_unit": None,
                }

        for col_idx, spec in charges_by_column.items():
            if col_idx >= len(row):
                continue
            value = to_primitive(row[col_idx])

            group_key = (spec["cost_name"], spec["container_type"], spec["charge_code"])
            if spec["metric"] == "currency":
                grouped[group_key]["currency"] = None if value in (None, "") else value
            elif spec["metric"] == "min":
                grouped[group_key]["flat_min"] = safe_float(value)
            else:
                grouped[group_key]["p_unit"] = safe_float(value)

        record["rates"] = [grouped[group_key] for group_key in ordered_cost_keys if group_key in grouped]
        records.append(record)

    return records, rate_card_validity


def build_output_path(source_file: Path):
    return OUTPUT_DIR / f"{source_file.stem}.json"


def main():
    files = list_input_files()
    selected_file = choose_input_file(files)
    records, rate_card_validity = parse_file(selected_file)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_file = build_output_path(selected_file)
    payload = {
        "schema_version": "1.0",
        "generated_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "source_file": str(selected_file),
        "source_sheet": TARGET_SHEET,
        "rate_card_validity": rate_card_validity,
        "record_count": len(records),
        "records": records,
    }
    output_file.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(f"Wrote {len(records)} records to {output_file}")


if __name__ == "__main__":
    main()
