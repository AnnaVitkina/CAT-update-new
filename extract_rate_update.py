import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
INPUT_DIR = ROOT / "input" / "rate updates"
OUTPUT_DIR = ROOT / "processing"


COMMON_FIELDS = [
    "KEY",
    "CARRIER",
    "CARRIERNAME",
    "SERVICE__C",
    "COMMODITY__C",
    "RATE_EFFECTIVE_DATE__C",
    "RATE_EXPIRATION_DATE__C",
    "SERVICE_GRADE_NUMBER__C",
    "SERVICE_GRADE_DEFINITION",
    "ORIGIN_REGION__C",
    "ORIGIN_LOCATION_NAME__C",
    "ORIGIN_COUNTRY__C",
    "ORIGIN_CITY__C",
    "DESTINATION_REGION__C",
    "DESTINATION_LOCATION_NAME__C",
    "DESTINATION_COUNTRY__C",
    "DESTINATION_CITY__C",
]

RATE_COL_RE = re.compile(
    r"^(?P<container>[0-9A-Z]+)_(?P<charge>[A-Z]+)_(?P<metric>CURRENCY|MIN|RATE)$"
)


def to_primitive(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    return value


def safe_float(value):
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def trim_service_for_transporeon(service_value):
    if service_value is None:
        return ""
    service = str(service_value)
    if service.startswith("OC_CNTR_"):
        service = service[len("OC_CNTR_") :]
    elif service.startswith("OC_CNTTR_"):
        service = service[len("OC_CNTTR_") :]
    if service.endswith("_BU"):
        service = service[: -len("_BU")]
    return service


def build_transporeon_id(row_map):
    carrier_code = "" if row_map.get("CARRIER") is None else str(row_map.get("CARRIER"))
    service = trim_service_for_transporeon(row_map.get("SERVICE__C"))
    origin = (
        ""
        if row_map.get("ORIGIN_LOCATION_NAME__C") is None
        else str(row_map.get("ORIGIN_LOCATION_NAME__C"))
    )
    destination = (
        ""
        if row_map.get("DESTINATION_LOCATION_NAME__C") is None
        else str(row_map.get("DESTINATION_LOCATION_NAME__C"))
    )
    return carrier_code + service + origin + destination


def list_input_files():
    files = sorted(INPUT_DIR.glob("*.xlsx"))
    if not files:
        raise FileNotFoundError(f"No .xlsx files found in {INPUT_DIR}")
    return files


def choose_input_file(files):
    if len(sys.argv) > 1:
        arg = sys.argv[1].strip()
        if arg.isdigit():
            idx = int(arg) - 1
            if 0 <= idx < len(files):
                return files[idx]
            raise ValueError(f"File index out of range: {arg}")
        candidate = INPUT_DIR / arg
        if candidate.exists():
            return candidate
        raise FileNotFoundError(f"Could not find file in input folder: {arg}")

    print("Choose a file to process:")
    for i, file_path in enumerate(files, start=1):
        print(f"{i}. {file_path.name}")

    selected = input("Enter file number: ").strip()
    if not selected.isdigit():
        raise ValueError("Please enter a valid number.")
    idx = int(selected) - 1
    if not (0 <= idx < len(files)):
        raise ValueError("Selected number is out of range.")
    return files[idx]


def parse_file(file_path: Path):
    workbook = load_workbook(file_path, data_only=True)
    records = []

    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue

        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        for row_index, row in enumerate(rows[1:], start=2):
            if all(cell is None or cell == "" for cell in row):
                continue

            row_map = dict(zip(headers, row))
            if row_map.get("KEY") in (None, ""):
                continue

            record = {
                "source_file": file_path.name,
                "sheet_name": sheet.title,
                "row_number": row_index,
                "route": {},
                "rates": [],
            }

            for col in COMMON_FIELDS:
                record["route"][col] = to_primitive(row_map.get(col))
            record["route"]["Transporeon ID"] = build_transporeon_id(row_map)

            grouped = {}
            for col, value in row_map.items():
                match = RATE_COL_RE.match(col)
                if not match:
                    continue

                container = match.group("container")
                charge = match.group("charge").lower()
                metric = match.group("metric").lower()

                grouped.setdefault(container, {}).setdefault(
                    charge, {"charge_code": charge}
                )[metric] = to_primitive(value)

            for container, charges in grouped.items():
                for charge_data in charges.values():
                    if (
                        charge_data.get("currency") is None
                        and charge_data.get("rate") is None
                        and charge_data.get("min") is None
                    ):
                        continue
                    record["rates"].append(
                        {
                            "container_type": container,
                            "charge_code": charge_data["charge_code"],
                            "currency": charge_data.get("currency"),
                            "min": safe_float(charge_data.get("min")),
                            "rate": safe_float(charge_data.get("rate")),
                        }
                    )

            records.append(record)

    return records


def build_output_path(source_file: Path):
    return OUTPUT_DIR / f"{source_file.stem}.json"


def main():
    files = list_input_files()
    selected_file = choose_input_file(files)
    records = parse_file(selected_file)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_file = build_output_path(selected_file)
    payload = {
        "schema_version": "1.0",
        "generated_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "source_file": str(selected_file),
        "record_count": len(records),
        "records": records,
    }
    output_file.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(f"Wrote {len(records)} records to {output_file}")


if __name__ == "__main__":
    main()
